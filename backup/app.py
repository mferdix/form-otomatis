from flask import Flask, render_template, request, redirect, url_for, send_file
import pandas as pd
from datetime import datetime
import os
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# Configuration
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
os.makedirs(DATA_DIR, exist_ok=True)
EXCEL_FILE = os.path.join(DATA_DIR, 'permintaan_server.xlsx')

def format_excel(filepath):
    """Format Excel file to be neat and professional"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # Set styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Format header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # Format all cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            if cell.value is None:
                cell.value = ""
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = min(adjusted_width, 30)
    
    wb.save(filepath)

def generate_server_name(platform, zone, name):
    """Generate server name based on rules"""
    if platform == "GCP":
        prefix = f"ZI{zone.upper()}"
    elif platform == "AWS":
        prefix = f"UI{zone.upper()}"
    else:
        prefix = ""
    return f"{prefix}{name}"

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    # Process form data
    platform = request.form.get('platform_type', '')
    zone = request.form.get('zone', '')
    server_name = request.form.get('server_name', '')
    
    # Generate full server name
    full_name = generate_server_name(platform, zone, server_name)

    # Prepare data for Excel
    form_data = {
        'Tanggal Input': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'Type': request.form.get('type', ''),
        'Property: Description of the Listed Property Information': request.form.get('project', ''),
        'No of Project': request.form.get('no_project', ''),
        'Number of CPUs': request.form.get('cpus', ''),
        'RAM (GB)': request.form.get('ram', ''),
        'Disk (GB)': request.form.get('disk', ''),
        'Zone': zone,
        'OS Platform': request.form.get('os_platform', ''),
        'Name': full_name,
        'Name Question': request.form.get('nama_pemohon', ''),
        'KIP': request.form.get('nip', ''),
        'Range: Examination': request.form.get('tanggal_permintaan', ''),
        'Email': request.form.get('email', ''),
        'Begins': request.form.get('bagian', ''),
        'Group: Sample Widths': request.form.get('group', ''),
        'Jangka Waktu': request.form.get('jangka_waktu', ''),
        'Perform Type': platform,
        'Environment': request.form.get('environment', '')
    }

    # Create new DataFrame
    df_new = pd.DataFrame([form_data])

    # Read existing data or create new file
    if os.path.exists(EXCEL_FILE):
        try:
            df_existing = pd.read_excel(EXCEL_FILE)
            df_combined = pd.concat([df_existing, df_new], ignore_index=True)
        except Exception as e:
            print(f"Error reading existing file: {e}")
            df_combined = df_new
    else:
        df_combined = df_new

    # Save to Excel
    df_combined.to_excel(EXCEL_FILE, index=False, engine='openpyxl')
    
    # Format the Excel file
    format_excel(EXCEL_FILE)
    
    return redirect(url_for('success'))

@app.route('/view-data')
def view_data():
    if not os.path.exists(EXCEL_FILE):
        return render_template('view_data.html', data=[], message="No data available")
    
    try:
        df = pd.read_excel(EXCEL_FILE)
        df = df.where(pd.notnull(df), None)  # Convert NaN to None
        
        # Define column order
        main_columns = [
            'Type',
            'Property: Description of the Listed Property Information',
            'No of Project',
            'Number of CPUs',
            'RAM (GB)',
            'Disk (GB)',
            'Zone',
            'OS Platform',
            'Name'
        ]
        
        # Get available columns
        available_columns = [col for col in main_columns if col in df.columns]
        data = df[available_columns].to_dict('records')
        
        return render_template('view_data.html', data=data, message=None)
    
    except Exception as e:
        return render_template('view_data.html', data=[], message=f"Error: {str(e)}")

@app.route('/download-excel')
def download_excel():
    if not os.path.exists(EXCEL_FILE):
        return redirect(url_for('view_data'))
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.read_excel(EXCEL_FILE).to_excel(writer, index=False)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='server_requests.xlsx'
    )

@app.route('/success')
def success():
    return render_template('success.html')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)