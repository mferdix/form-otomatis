from flask import Flask, render_template, request, redirect, url_for, send_file
import pandas as pd
from datetime import datetime
import os
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# ========== KONFIGURASI ==========
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
os.makedirs(DATA_DIR, exist_ok=True)
EXCEL_FILE = os.path.join(DATA_DIR, 'permintaan_server.xlsx')

# ========== FUNGSI UTILITAS ==========
def format_excel(filepath):
    """Format file Excel agar rapi"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # Style untuk header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Format header
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # Format semua cell
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            if cell.value is None:
                cell.value = ""
    
    # Auto adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = min(adjusted_width, 30)
    
    wb.save(filepath)

def create_sample_data():
    """Buat data contoh jika file Excel belum ada"""
    sample_data = {
        'Tanggal Input': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        'Nama Pemohon': ['CONTOH DATA'],
        'NIP': ['123456'],
        'Tanggal Permintaan': [datetime.now().strftime('%Y-%m-%d')],
        'Email': ['contoh@email.com'],
        'Bagian': ['IT'],
        'Group': ['Development'],
        'Jangka Waktu': ['1 Tahun'],
        'Platform Type': ['GCP'],
        'Environment': ['Development'],
        'Type': ['Virtual Machine'],
        'Project/Aplikasi/Perangkat': ['Aplikasi Internal'],
        'No of Project/Aplikasi/Perangkat': ['PRJ-001'],
        'Number of CPUs': [4],
        'RAM (GB)': [8],
        'Disk (GB)': [50],
        'Zone': ['A'],
        'OS Platform': ['Linux'],
        'Cloud Zone Prefix': ['ZIA'],
        'Server Name Input': ['APPSERVER'],
        'Server Name': ['ZIAAPPSERVER']
    }
    df = pd.DataFrame(sample_data)
    df.to_excel(EXCEL_FILE, index=False)
    format_excel(EXCEL_FILE)

# ========== ROUTES ==========
@app.route('/')
def home():
    """Redirect langsung ke view data"""
    return redirect(url_for('view_data'))

@app.route('/form')
def form():
    """Tampilkan form input"""
    return render_template('form.html', tanggal_hari_ini=datetime.now().strftime('%Y-%m-%d'))

@app.route('/submit', methods=['POST'])
def submit():
    """Proses form submit"""
    # Generate Cloud Zone Prefix
    platform = request.form.get('platform_type')
    zone = request.form.get('zone')
    cloud_zone = ""
    
    if platform == "GCP":
        if zone == "A": cloud_zone = "ZIA"
        elif zone == "B": cloud_zone = "ZIB"
        elif zone == "C": cloud_zone = "ZIC"
    elif platform == "AWS":
        if zone == "A": cloud_zone = "UIA"
        elif zone == "B": cloud_zone = "UIB"
        elif zone == "C": cloud_zone = "UIC"
    
    # Gabungkan Cloud Zone + Server Name
    server_name = request.form.get('server_name')
    full_server_name = f"{server_name}"

    # Siapkan data untuk Excel
    form_data = {
        'Tanggal Input': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'Nama Pemohon': request.form.get('nama_pemohon'),
        'NIP': request.form.get('nip'),
        'Tanggal Permintaan': request.form.get('tanggal_permintaan'),
        'Email': request.form.get('email'),
        'Bagian': request.form.get('bagian'),
        'Group': request.form.get('group'),
        'Jangka Waktu': request.form.get('jangka_waktu'),
        'Platform Type': platform,
        'Environment': request.form.get('environment'),
        'Type': request.form.get('type'),
        # 'Project/Aplikasi/Perangkat': request.form.get('project'),
        # 'No of Project/Aplikasi/Perangkat': request.form.get('no_project'),
        'Number of CPUs': request.form.get('cpus'),
        'RAM (GB)': request.form.get('ram'),
        'Disk (GB)': request.form.get('disk'),
        'Zone': zone,
        'OS Platform': request.form.get('os_platform'),
        'Cloud Zone Prefix': cloud_zone,
        # 'Server Name Input': server_name,
        'Server Name': full_server_name
    }

    # Simpan ke Excel
    df_new = pd.DataFrame([form_data])
    if os.path.exists(EXCEL_FILE):
        df_existing = pd.read_excel(EXCEL_FILE)
        df_combined = pd.concat([df_existing, df_new], ignore_index=True)
    else:
        df_combined = df_new
    
    df_combined.to_excel(EXCEL_FILE, index=False)
    format_excel(EXCEL_FILE)
    
    return redirect(url_for('view_data'))

@app.route('/view-data')
def view_data():
    """Tampilkan semua data dengan semua kolom"""
    if not os.path.exists(EXCEL_FILE):
        create_sample_data()
    
    try:
        df = pd.read_excel(EXCEL_FILE)
        df = df.fillna('')
        
        # Konversi semua kolom ke string untuk tampilan
        for col in df.columns:
            df[col] = df[col].astype(str)
        
        # Urutkan kolom secara logis
        column_order = [
            'Tanggal Input',
            'Nama Pemohon',
            'NIP',
            'Tanggal Permintaan',
            'Email',
            'Bagian',
            'Group',
            'Jangka Waktu',
            'Platform Type',
            'Environment',
            'Type',
            # 'Project/Aplikasi/Perangkat',
            # 'No of Project/Aplikasi/Perangkat',
            'Number of CPUs',
            'RAM (GB)',
            'Disk (GB)',
            'Zone',
            'OS Platform',
            'Cloud Zone Prefix',
            # 'Server Name Input',
            'Server Name'
        ]
        
        # Filter hanya kolom yang ada di DataFrame
        available_columns = [col for col in column_order if col in df.columns]
        df = df[available_columns]
        
        data = df.to_dict('records')
        return render_template('view_data.html', data=data, columns=available_columns)
    
    except Exception as e:
        return render_template('error.html', message=f"Error membaca data: {str(e)}")

@app.route('/download')
def download():
    """Download file Excel"""
    if not os.path.exists(EXCEL_FILE):
        return redirect(url_for('view_data'))
    
    return send_file(
        EXCEL_FILE,
        as_attachment=True,
        download_name="permintaan_server.xlsx"
    )

if __name__ == '__main__':
    app.run(debug=True)